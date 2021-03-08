# author:       Alexandru Boldea
# date:         Oct 2020

# Code structure as follows:
#-----------------------------
# import modules
# window layout definition
# binary tree definition
# excel list function
# get working folder from the user
# create window
# loop through events in the window
# close window

# to create ".exe" file, the console command line is: pyinstaller --onefile -w -F DuplicateSongFinder.py

import sys
import os # needed to get the info about the files and folders (path, size) and also to delete (unlink) the files
import re # needed to use Regex to get the file extensions
from win32api import GetSystemMetrics # this is needed to get the screen resolution and to position the app window relative to the center of the screen
import PySimpleGUI as sg # this is the wonderful GUI
import openpyxl # needed to generate the song list in xlsx format
from openpyxl.styles import Font # needed for auto resize of columns

#----------------------------This is how the window looks----------------------------#

sg.theme('DarkGrey6')   # Color theme

sg.popup_ok('Thank you for using this app! \n\nBefore we start, please read this info carefully.\n\nThis app will permanently delete files from your computer,\nso '
            'think 3 times before making your choice.\n\nIt has been developed for audio files, it will automatically delete text, database and playlist files without asking.\n\n'
            'You can stop the process at any time by closing the app, then pick up where you left off.\n\n'
            'Let\'s get started by choosing the folder you want to clean.\nClick \'OK\' to start.\n', no_titlebar=True, grab_anywhere=True)


col_layout_1 = [[sg.Text(''), sg.Text(size=(30, 7), key='-LOC1-')],
                []]

col_layout_2 = [[sg.Text(''), sg.Text(size=(30, 7), key='-LOC2-')],
                []]

layout = [  [sg.Text('Selected folder is: ')],                          # everything in the main window is an element in this list of lists
            [sg.Text(size=(60,1), key='-FOLDER-')],
            [sg.Button('Start')],
            [sg.Text('File to delete: '),
             sg.Text(size=(50, 1), font = 'Any 10', text_color='yellow', key='-SONGNAME-')],
            [sg.Frame('First location: ', col_layout_1, element_justification='c'),
             sg.VSeparator(),
             sg.Frame('Second location: ', col_layout_2, element_justification='c')],
            [sg.Text('Delete file from:')],
            [sg.Text(' '), sg.Text(size=(9, 1)), sg.Button('Location 1'), sg.Text(''), sg.Text(size=(9, 1)),
             sg.VSeparator(),
             sg.Text(''), sg.Text(size=(9, 1)),sg.Button('Location 2')],
            [sg.Text(' '), sg.Text(size=(1, 1))],
            [sg.Text(' '), sg.Text(size=(28, 1)), sg.Button('Skip file')],
            [sg.Text(' '), sg.Text(size=(1, 1))],
            [sg.Button('Generate Excel list of all the files')],
            [sg.Text('Proudly done using PySimpleGUI', size=(70, 1), justification='r', text_color='Grey')]] # Huge "thank you" to Mike for making this GUI resource available for free

#--------------------------Define the binary tree Node class---------------------------#

class Node:

    def __init__(self, data, path, size):

        self.left = None
        self.right = None
        self.data = data
        self.hdata = hash(data) # this property is an unique integer used to build the binary tree
        self.path = path # needed for the user info
        self.size = int(size) # needed for the user info

    def insert(self, data, path, size):
        # Compare the new value with the parent node
        fileExtensionRegex = re.compile(r'.\w+$')
        mo = fileExtensionRegex.search(data)
        if mo.group() in ('.db', '.txt', '.m3u', '.nfo', '.torrent', '.sfv', '.log', '.pls', '.doc',
                          '.html', '.ini', '.TXT', '.url', '.htm', '.m3u8', '.cue', '.BUP', '.IFO',
                          '.tmp', '.missing', '.rtf', '.DS_Store', '.xlsx'):
            os.unlink(path + '/' + data)
        if mo.group() in ('.mp3', '.MP3', '.wma', '.Mp3', '.m4a', '.flac', '.mpg',
                          '.mp4', '.VOB', '.wav', '.mkv', '.avi'):
            if self.hdata and self.size:
                if (hash(data) + size) < (self.hdata + self.size):
                    if self.left is None:
                        self.left = Node(data, path, size)
                    else:
                        self.left.insert(data, path, size)
                elif (hash(data) + size) > (self.hdata + self.size):
                    if self.right is None:
                        self.right = Node(data, path, size)
                    else:
                        self.right.insert(data, path, size)
                else:                                       # this is where the file actually gets deleted
                    displayPath1 = str.replace(path, workingFolder, '')
                    displayPath2 = str.replace(self.path, workingFolder, '')
                    displaySize1 = '{:.2f}'.format(size/1048576)
                    displaySize2 = '{:.2f}'.format(self.size/1048576)
                    window['-SONGNAME-'].update(data)                   # the song name is displayed in the app window
                    window['-LOC1-'].update('\n' + displayPath1 + '\n\nSize: ' + displaySize1 + ' MB')   # display the first location and size
                    window['-LOC2-'].update('\n' + displayPath2 + '\n\nSize: ' + displaySize2 + ' MB')   # display the second location and size
                    event, values = window.read()                       # to be able to get window events related to the "insert" method, this line needs to be inside the method
                    if event == sg.WIN_CLOSED:              # When this event is not included, an error occurs when closing the window while running ("trying to read a closed window")
                        sys.exit()
                    if event == 'Location 1':               # this is the "click" event on the button "Location 1"
                        os.unlink(path + data)
                    if event == 'Location 2':               # this is the "click" event on the button "Location 2"
                        os.unlink(self.path + self.data)
                        self.data = data
                        self.path = path
                        self.size = size
                    if event == 'Skip file':                # this is the "click" event on the button "Skip file", it leaves both files on the drive
                        return None
            else:
                self.data = data
                self.path = path
                self.size = size

#----------------------------Define the excel list function----------------------------#

def ListGenerator(workingFolder):
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    sheet.title = 'All'
    sheet['A1'] = 'Song Title'
    sheet['B1'] = 'Song Size'
    sheet['C1'] = 'Song Path'
    title_font = Font(color='00800000', bold=True, size=15)
    sheet.row_dimensions[1].height = 20
    sheet['A1'].font = title_font
    sheet['B1'].font = title_font
    sheet['C1'].font = title_font
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 20
    line = 2
    for folderName, subfolders, filenames in os.walk(workingFolder):
        for item in filenames:
            itemPath = os.path.join(folderName + '/')
            itemSize = '{:.2f}'.format(int(os.stat(os.path.join(folderName + '/' + item)).st_size)/1048576) + ' MB'
            name = sheet.cell(row=line, column=1, value=item)
            size = sheet.cell(row=line, column=2, value=itemSize)
            path = sheet.cell(row=line, column=3, value=itemPath)
            line += 1
    os.chdir(workingFolder)
    wb.save('SongList-Generated.xlsx')
    sg.popup_ok('Done!\n'+str(line-2)+' files added to list.', no_titlebar=True, grab_anywhere=True,
                text_color='black', background_color='DarkGrey')

#--------------------------------Get the working folder--------------------------------#

workingFolder = sg.popup_get_folder('Select the folder to be searched for duplicates:', no_titlebar=True, grab_anywhere=True,
                                    location=(GetSystemMetrics(0)/2-100, GetSystemMetrics(1)/2-50))

#----------Create the Window and position it relative to the screen resolution---------#

window = sg.Window('Duplicate Song Finder', layout, keep_on_top=False, alpha_channel=1,
                   location=(GetSystemMetrics(0)/2-250, GetSystemMetrics(1)/2-150))

#----------Event Loop to process "events" and get the "values" of the inputs-----------#

while True:
    event, values = window.read(timeout=200)
    # print(event, values ) # to be deleted
    window['-FOLDER-'].update(workingFolder)
    if event == sg.TIMEOUT_KEY:
        continue
    if event == sg.WIN_CLOSED: # if user closes window
        sys.exit()
    if event == 'Start':
        window.refresh()
        window['Start'].update(disabled=True)
        if workingFolder:
            root = Node('a', 'b', 0)
            for folderName, subf, filenames in os.walk(workingFolder, topdown=False):
                for data in filenames:
                    path = os.path.join(folderName + '/')
                    size = os.stat(os.path.join(folderName + '/' + data)).st_size
                    window['Generate Excel list of all the files'].update(disabled=True)
                    root.insert(data, path, size)
                    if not os.listdir(folderName):
                        os.rmdir(folderName)
                window['Generate Excel list of all the files'].update(disabled=False)
            window['-SONGNAME-'].update('All finished!')
            window['-LOC1-'].update('')
            window['-LOC2-'].update('')
        else:
            sg.popup_error('No folder provided', no_titlebar=True, grab_anywhere=True, text_color='black',
                           background_color='DarkGrey')
        if event == sg.WIN_CLOSED:
                sys.exit()
    if event == 'Generate Excel list of all the files':
        try:
            ListGenerator(workingFolder)
        except:
            sg.popup_error('No folder provided', no_titlebar=True, grab_anywhere=True, text_color='black',
                           background_color='DarkGrey')

#----------------------------------closes window--------------------------------------#

window.close()
