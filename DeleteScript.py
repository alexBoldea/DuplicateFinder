#! python3

# FileNotFoundError: [WinError 2] The system cannot find the file specified - FIXED by "os.path.exists()"
# ===========================================================================================
# This happens when the file appears in the list on a line after it was already deleted above

import re
import os
import openpyxl

songNameRegex = re.compile(r'.\w+$')
def deleteSong(songName, firstPath, secondPath):
    mo = songNameRegex.search(songName)
    if mo.group() in ('.db', '.txt', '.m3u', '.nfo', '.torrent', '.sfv', '.log', '.pls', '.doc',
                      '.html', '.ini', '.TXT', '.url', '.htm', '.m3u8', '.cue', '.BUP', '.IFO',
                      '.tmp', '.missing', '.rtf', '.DS_Store'):
        #os.unlink(songName)
        print(songName)
    print('Option 1: ', firstPath, songName)
    print('Option 2: ', secondPath, songName)
    choice = input("Choose which song to delete: 1 / 2")
    if choice in ('1', '2'):
        if choice == '1':
            if os.path.exists(firstPath + songName):
                os.unlink(firstPath + songName)
                #print('DELETE: ', firstPath, songName)
            else:
                print ('File does not exist')
        elif choice == '2':
            if os.path.exists(secondPath + songName):
                os.unlink(secondPath + songName)
                #print('DELETE: ', secondPath, songName)
            else:
                print('File does not exist')
        return 1
    else:
        print('Invalid selection, try again!')
        return 0

# Pass through excel song list duplicates, delete the needed files

os.chdir('c:\\users\\alex_\\Desktop\\dis')
path = "c:\\users\\alex_\\Desktop\\dis\\SongListTest.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj["Duplicates"]

for i in range (2, 32):
    name = sheet_obj.cell(row = i, column = 1).value
    path1 = sheet_obj.cell(row = i, column = 3).value
    path2 = sheet_obj.cell(row = i, column = 4).value
    while not deleteSong(name, path1, path2):
        deleteSong(name, path1, path2)
