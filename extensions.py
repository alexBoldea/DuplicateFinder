#! python3

import openpyxl
import os
import re

songNameRegex = re.compile(r'.\w+$')                                            # regex to find extension goes here
active = re.compile(r'.\S+')                                                    # regex to get the list of extensions

os.chdir('c:\\users\\alex_\\Desktop')
path = "c:\\users\\alex_\\Desktop\\SongList.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active                                                       # selected the active sheet from the excel file

inputFile = open('C:\\Users\\alex_\\Desktop\\extensions.txt', 'w')
init = songNameRegex.search(sheet_obj.cell(row = 2, column = 1).value)
ext = init.group()
inputFile.write(' ' + ext + ' ')
inputFile.close()                                                               # wrote the first extension in the text file


for i in range (3, 36369):
    mo = songNameRegex.search(sheet_obj.cell(row = i, column = 1).value)
    inputFile = open('C:\\Users\\alex_\\Desktop\\extensions.txt', 'a')
    inputFile.write(mo.group() + ' ')    
    inputFile.close()                                                           # wrote all extensions in the file

inputFile = open('C:\\Users\\alex_\\Desktop\\extensions.txt')
boo = active.findall(inputFile.read())
#print(boo)

boo = list(dict.fromkeys(boo))                                                  # remove duplicates
#print(boo)
inputFile.close()

inputFile = open('C:\\Users\\alex_\\Desktop\\extensions.txt', 'w')
inputFile.write('Extensions: \n')
inputFile.close()
for i in range (len(boo)):
    inputFile = open('C:\\Users\\alex_\\Desktop\\extensions.txt', 'a')
    inputFile.write(boo[i] + '\n')    
    inputFile.close()
                                                                                # output is a list with all different file extensions
